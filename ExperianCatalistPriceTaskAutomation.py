import os

#find the cwd (current working directory); location where downloaded excel sheet will be placed
cwd = os.getcwd()

#function to download email attachment including success or failure console message
def downloadAttachment():
    import win32com.client
    from datetime import datetime
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)
    messages = inbox.Items
    
    def confirmation(): 
        for message in messages:
            if message.Subject == "FW: Experian Catalist Price Averages" and message.sentOn.date() == datetime.now().date():
                message.Attachments.Item(1).SaveAsFile(cwd+ "\\ExperianDailyAverage.xlsx")
                return True

    if confirmation() == True:
        print("Email found and attachment downloaded to " + cwd)
    else:
        print("Experian Catalist Price Averages email cannot be found.")    

#file downloader and puts into the cwd - must go before path declarations
#downloadAttachment()    
shareDrivePath = "//LS-WTGL03A//share//MV"
worksheet1 = cwd + "\\ExperianDailyAverage.xlsx"
worksheet2 = shareDrivePath + "//Pump Prices vs Platts.xlsx"

#function to delete empty rows in specified sheet
def deleteEmptyRowsInCertainSheet(worksheet):
    import openpyxl
    wb = openpyxl.load_workbook(worksheet)
    ws = wb["Imports"]
    maxRow = ws.max_row
    print("There are currently " + str(maxRow) + " rows of data in the sheet.")
    emptyRows = []
    for row in ws.iter_rows():
        if row[0].value == None:
            emptyRows.append(row)
    print(str(emptyRows.__len__()) + " empty rows found.")
    #will only begin deleting if there is more than 1 empty row, since if there is not, then there is no empty rows.
    if emptyRows.__len__() > 0:
        for row in emptyRows:
            ws.delete_rows(row[0].row)
        print("Empty rows deleted.")    
        wb.save(worksheet)

#function that moves selected rows from one sheet to another
def rowMover(worksheet1, worksheet2):
    import openpyxl
    import datetime
    print("Beginning to move data from downloaded Excel sheet to existing Excel sheet.")
    book1 = openpyxl.load_workbook(worksheet1)
    sheet1 = book1.active
    book2 = openpyxl.load_workbook(worksheet2)
    sheet2 = book2["Imports"]
    #data comes in 3 rows, first row being header, so must copy over 2nd and 3rd row, but must validate that they do not exist already.
    #sometimes experian sends the same day's data twice.

    #must first format experian data to be in datetime format as we cannot compare the dates already in excel to the ones sent by experian
    sheet1_a2Raw = sheet1["A2"].value
    sheet1_a2Formatted = datetime.datetime.strptime(sheet1_a2Raw, "%d/%m/%Y")

    if sheet1_a2Formatted == sheet2['A' + str(sheet2.max_row-1)].value or sheet1_a2Formatted == sheet2['A' + str(sheet2.max_row)].value:
        print("Experian Catalist Price Averages for day " + sheet1['A2'].value + " already exists in destination sheet.")
    else:
        sheet1Row2 = sheet1.iter_rows(min_row=2, max_row=2, min_col=1, max_col=sheet1.max_column, values_only=True)
        for cell in sheet1Row2:
            sheet2.append(cell)
            book2.save(worksheet2)
        print("Added Experian Catalist Price Averages for day " + sheet1['A2'].value + " to " + worksheet2)

    sheet1_a3Raw = sheet1["A3"].value
    sheet1_a3Formatted = datetime.datetime.strptime(sheet1_a3Raw, "%d/%m/%Y")

    if sheet1_a3Formatted == sheet2['A' + str(sheet2.max_row-1)].value or sheet1_a3Formatted == sheet2['A' + str(sheet2.max_row)].value:
        print("Experian Catalist Price Averages for day " + sheet1['A3'].value + " already exists in destination sheet.")
    else:
        sheet1Row3 = sheet1.iter_rows(min_row=3, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column, values_only=True)
        for cell in sheet1Row3:
            sheet2.append(cell)
            book2.save(worksheet2)
        print("Added Experian Catalist Price Averages for day " + sheet1['A3'].value + " to " + worksheet2) 

#function for formatting the new cells to be in correct format(date/number/alignment)
def cellFormatting(worksheet):
    import openpyxl
    import datetime
    print("Beginning formatting newly entered cells to work in formulas on other sheets.")
    book = openpyxl.load_workbook(worksheet)
    sheet = book["Imports"]
    maxrow = sheet.max_row

    #formatting for date cells
    def excelDateToNumber(worksheet, cell):
        try:
            dateText = sheet[cell].value
            #print(sheet[cell].value)
            #print(dateText)
            #convert dateText to datetime object
            date = datetime.datetime.strptime(dateText, "%d/%m/%Y")
            #print(date)
            sheet[cell].value = date
            sheet[cell].number_format = 'dd/mm/yyyy'
            #print(sheet[cell].value)
            print(cell + " converted to datetime format")
            
        except:
            print(cell + " is already in datetime format")
        book.save(worksheet)

    excelDateToNumber(worksheet, "A" + str(maxrow-1))
    excelDateToNumber(worksheet, "A" + str(maxrow))

    
    #formatting for number cells in first row of newly inputted data
    for c in sheet.iter_cols(min_row=maxrow-1, max_row=maxrow, min_col=2, max_col=5):
        c[0].number_format = '#,##0.00'
        c[0].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='bottom')
    #formatting for number cells in second row of newly inputted data
    for c in sheet.iter_cols(min_row=maxrow, max_row=maxrow, min_col=2, max_col=5):
        c[0].number_format = '#,##0.00'
        c[0].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='bottom') 
    book.save(worksheet)
    print("Cell formatting complete to match destination.")

#function to make sure downloaded data is in expected format, if not, program will not run
def experianCatalistPriceTaskAutomation(worksheet1, worksheet2):
    import openpyxl
    book = openpyxl.load_workbook(worksheet1)
    sheet = book.active
    maxrow = sheet.max_row

    #formatting for date cells
    if maxrow == 3:
        print("Downloaded data is in expected format, beginning program")
        deleteEmptyRowsInCertainSheet(worksheet2)
        rowMover(worksheet1, worksheet2)
        cellFormatting(worksheet2)
        print("Program has completed.")
    else:
        print("Downloaded data is not in expected format, program will not run.")

experianCatalistPriceTaskAutomation(worksheet1, worksheet2)
